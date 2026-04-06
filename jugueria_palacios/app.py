from flask import Flask, render_template, request, jsonify, send_file
import sqlite3, openpyxl, os, io, json
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta

app = Flask(__name__)
DB_PATH = os.path.join(os.path.dirname(__file__), 'palacios.db')

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS ambientes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT UNIQUE
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS mesas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        numero INTEGER, ambiente_id INTEGER DEFAULT 1,
        estado TEXT DEFAULT 'libre',
        UNIQUE(numero, ambiente_id)
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS pedidos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha TEXT, hora TEXT,
        mesa TEXT, ambiente TEXT,
        tipo TEXT DEFAULT 'mesa',
        estado TEXT DEFAULT 'abierto',
        metodo_pago TEXT DEFAULT 'efectivo',
        total REAL DEFAULT 0
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS pedido_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        pedido_id INTEGER,
        categoria TEXT, producto TEXT, tamano TEXT,
        precio REAL, personalizado TEXT, qty INTEGER DEFAULT 1,
        estado TEXT DEFAULT 'preparando',
        FOREIGN KEY(pedido_id) REFERENCES pedidos(id)
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS ventas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha TEXT, hora TEXT, mesa TEXT,
        categoria TEXT, producto TEXT, tamano TEXT,
        precio REAL, personalizado TEXT, metodo_pago TEXT DEFAULT 'efectivo'
    )''')
    # Ambiente por defecto
    c.execute("INSERT OR IGNORE INTO ambientes (id, nombre) VALUES (1, 'Principal')")
    c.execute("INSERT OR IGNORE INTO ambientes (id, nombre) VALUES (2, 'Terraza')")
    # Mesas por defecto
    for i in range(1, 13):
        c.execute('INSERT OR IGNORE INTO mesas (numero, ambiente_id, estado) VALUES (?, 1, "libre")', (i,))
    for i in range(1, 7):
        c.execute('INSERT OR IGNORE INTO mesas (numero, ambiente_id, estado) VALUES (?, 2, "libre")', (i,))
    # Migrar columna mesa en ventas si no existe
    try: c.execute('ALTER TABLE ventas ADD COLUMN mesa TEXT DEFAULT ""')
    except: pass
    conn.commit()
    conn.close()

init_db()

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/manifest.json')
def manifest():
    return send_file(os.path.join(os.path.dirname(__file__), 'static', 'manifest.json'), mimetype='application/manifest+json')

# ── AMBIENTES ──
@app.route('/api/ambientes')
def get_ambientes():
    conn = get_db()
    rows = conn.execute('SELECT * FROM ambientes ORDER BY id').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/ambientes', methods=['POST'])
def crear_ambiente():
    nombre = request.json.get('nombre', '')
    conn = get_db()
    conn.execute('INSERT INTO ambientes (nombre) VALUES (?)', (nombre,))
    # Crear 6 mesas en el nuevo ambiente
    amb_id = conn.execute('SELECT id FROM ambientes WHERE nombre=?', (nombre,)).fetchone()['id']
    for i in range(1, 7):
        conn.execute('INSERT OR IGNORE INTO mesas (numero, ambiente_id) VALUES (?, ?)', (i, amb_id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ── MESAS ──
@app.route('/api/mesas')
def get_mesas():
    amb_id = request.args.get('ambiente_id', 1)
    conn = get_db()
    mesas = conn.execute('SELECT * FROM mesas WHERE ambiente_id=? ORDER BY numero', (amb_id,)).fetchall()
    conn.close()
    return jsonify([dict(m) for m in mesas])

@app.route('/api/mesa/<int:mesa_id>/estado', methods=['POST'])
def set_mesa_estado(mesa_id):
    estado = request.json.get('estado', 'libre')
    conn = get_db()
    conn.execute('UPDATE mesas SET estado=? WHERE id=?', (estado, mesa_id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ── PEDIDOS ──
@app.route('/api/pedido', methods=['POST'])
def crear_pedido():
    data = request.json
    now = datetime.now()
    conn = get_db()
    cur = conn.execute('''INSERT INTO pedidos (fecha, hora, mesa, ambiente, tipo, estado, metodo_pago)
        VALUES (?,?,?,?,?,?,?)''', (
        now.strftime('%Y-%m-%d'), now.strftime('%H:%M:%S'),
        data.get('mesa',''), data.get('ambiente',''),
        data.get('tipo','mesa'), 'abierto',
        data.get('metodo_pago','efectivo')
    ))
    pedido_id = cur.lastrowid
    for item in data.get('items', []):
        conn.execute('''INSERT INTO pedido_items (pedido_id, categoria, producto, tamano, precio, personalizado, qty)
            VALUES (?,?,?,?,?,?,?)''', (
            pedido_id, item['categoria'], item['producto'],
            item.get('tamano',''), item['precio'],
            item.get('personalizado',''), item.get('qty',1)
        ))
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'pedido_id': pedido_id})

@app.route('/api/pedido/<int:pedido_id>/cerrar', methods=['POST'])
def cerrar_pedido(pedido_id):
    metodo = request.json.get('metodo_pago', 'efectivo')
    conn = get_db()
    items = conn.execute('SELECT * FROM pedido_items WHERE pedido_id=?', (pedido_id,)).fetchall()
    ped = conn.execute('SELECT * FROM pedidos WHERE id=?', (pedido_id,)).fetchone()
    total = sum(i['precio'] * i['qty'] for i in items)
    now = datetime.now()
    # Registrar en ventas
    for item in items:
        for _ in range(item['qty']):
            conn.execute('''INSERT INTO ventas (fecha, hora, mesa, categoria, producto, tamano, precio, personalizado, metodo_pago)
                VALUES (?,?,?,?,?,?,?,?,?)''', (
                now.strftime('%Y-%m-%d'), now.strftime('%H:%M:%S'),
                ped['mesa'], item['categoria'], item['producto'],
                item['tamano'], item['precio'], item['personalizado'], metodo
            ))
    conn.execute('UPDATE pedidos SET estado="cerrado", metodo_pago=?, total=? WHERE id=?', (metodo, total, pedido_id))
    # Liberar mesa
    if ped['mesa']:
        conn.execute('UPDATE mesas SET estado="libre" WHERE numero=? AND ambiente_id=(SELECT id FROM ambientes WHERE nombre=?)',
                     (ped['mesa'], ped['ambiente']))
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'total': total})

@app.route('/api/pedidos/activos')
def pedidos_activos():
    conn = get_db()
    rows = conn.execute('''SELECT p.*, 
        (SELECT SUM(precio*qty) FROM pedido_items WHERE pedido_id=p.id) as total
        FROM pedidos p WHERE p.estado="abierto" ORDER BY p.hora DESC''').fetchall()
    result = []
    for r in rows:
        d = dict(r)
        items = conn.execute('SELECT * FROM pedido_items WHERE pedido_id=?', (r['id'],)).fetchall()
        d['items'] = [dict(i) for i in items]
        result.append(d)
    conn.close()
    return jsonify(result)

@app.route('/api/pedido/<int:pedido_id>/item/<int:item_id>/estado', methods=['POST'])
def update_item_estado(pedido_id, item_id):
    estado = request.json.get('estado')
    conn = get_db()
    conn.execute('UPDATE pedido_items SET estado=? WHERE id=? AND pedido_id=?', (estado, item_id, pedido_id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ── VENTAS (legacy) ──
@app.route('/api/venta', methods=['POST'])
def registrar_venta():
    data = request.json
    conn = get_db()
    now = datetime.now()
    conn.execute('''INSERT INTO ventas (fecha,hora,mesa,categoria,producto,tamano,precio,personalizado,metodo_pago)
        VALUES (?,?,?,?,?,?,?,?,?)''', (
        now.strftime('%Y-%m-%d'), now.strftime('%H:%M:%S'),
        data.get('mesa',''), data.get('categoria',''), data.get('producto',''),
        data.get('tamano',''), data.get('precio',0), data.get('personalizado',''),
        data.get('metodo_pago','efectivo')
    ))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/ventas/hoy')
def ventas_hoy():
    conn = get_db()
    hoy = datetime.now().strftime('%Y-%m-%d')
    rows = conn.execute('SELECT * FROM ventas WHERE fecha=? ORDER BY hora DESC', (hoy,)).fetchall()
    total = conn.execute('SELECT SUM(precio) as t FROM ventas WHERE fecha=?', (hoy,)).fetchone()['t'] or 0
    conn.close()
    return jsonify({'ventas': [dict(r) for r in rows], 'total': total})

# ── REPORTES ──
@app.route('/api/reporte')
def reporte():
    hoy = datetime.now()
    inicio = request.args.get('inicio', hoy.strftime('%Y-%m-%d'))
    fin = request.args.get('fin', hoy.strftime('%Y-%m-%d'))
    conn = get_db()
    rows = conn.execute('SELECT * FROM ventas WHERE fecha BETWEEN ? AND ? ORDER BY fecha,hora', (inicio,fin)).fetchall()
    resumen = conn.execute('''SELECT producto,tamano,categoria,COUNT(*) as cantidad,SUM(precio) as total
        FROM ventas WHERE fecha BETWEEN ? AND ? GROUP BY producto,tamano ORDER BY total DESC''', (inicio,fin)).fetchall()
    total_general = conn.execute('SELECT SUM(precio) as t FROM ventas WHERE fecha BETWEEN ? AND ?', (inicio,fin)).fetchone()['t'] or 0
    por_dia = conn.execute('''SELECT fecha,SUM(precio) as total,COUNT(*) as cantidad
        FROM ventas WHERE fecha BETWEEN ? AND ? GROUP BY fecha ORDER BY fecha''', (inicio,fin)).fetchall()
    por_mesa = conn.execute('''SELECT mesa,COUNT(*) as pedidos,SUM(precio) as total
        FROM ventas WHERE fecha BETWEEN ? AND ? AND mesa!='' GROUP BY mesa ORDER BY total DESC''', (inicio,fin)).fetchall()
    mesas_unicas = conn.execute(
        "SELECT DISTINCT mesa FROM ventas WHERE fecha BETWEEN ? AND ? AND mesa!='' ORDER BY mesa",
        (inicio,fin)).fetchall()

    wb = openpyxl.Workbook()
    naranja='E65100'; verde='1B5E20'

    def hcell(ws,r,c,v,bg=verde,fg='FFFFFF',sz=10,bold=True):
        cell=ws.cell(row=r,column=c,value=v)
        cell.font=Font(bold=bold,color=fg,size=sz)
        cell.fill=PatternFill('solid',fgColor=bg)
        cell.alignment=Alignment(horizontal='center',vertical='center')
        thin=Side(style='thin',color='CCCCCC')
        cell.border=Border(left=thin,right=thin,top=thin,bottom=thin)
        return cell

    def dcell(ws,r,c,v,bg='FFFFFF',bold=False,fmt=None,align='center'):
        cell=ws.cell(row=r,column=c,value=v)
        cell.font=Font(bold=bold,size=10)
        cell.fill=PatternFill('solid',fgColor=bg)
        cell.alignment=Alignment(horizontal=align,vertical='center')
        if fmt: cell.number_format=fmt
        thin=Side(style='thin',color='EEEEEE')
        cell.border=Border(left=thin,right=thin,top=thin,bottom=thin)
        return cell

    # Hoja resumen
    ws1=wb.active; ws1.title='Resumen'
    ws1.merge_cells('A1:F1'); t=ws1['A1']
    t.value=f'JUGUERÍA PALACIOS — {inicio} al {fin}'
    t.font=Font(bold=True,size=13,color='FFFFFF'); t.fill=PatternFill('solid',fgColor=naranja)
    t.alignment=Alignment(horizontal='center',vertical='center'); ws1.row_dimensions[1].height=28
    ws1.merge_cells('A2:F2'); t2=ws1['A2']; t2.value=f'TOTAL: S/ {total_general:.2f}'
    t2.font=Font(bold=True,size=12,color=verde); t2.alignment=Alignment(horizontal='center')
    for i,h in enumerate(['Producto','Tamaño','Categoría','Cant.','Total (S/)'],1): hcell(ws1,3,i,h)
    for ri,row in enumerate(resumen,4):
        bg='FFF9C4' if ri%2==0 else 'FFFFFF'
        dcell(ws1,ri,1,row['producto'],bg,align='left'); dcell(ws1,ri,2,row['tamano'] or '-',bg)
        dcell(ws1,ri,3,row['categoria'],bg); dcell(ws1,ri,4,row['cantidad'],bg,True)
        dcell(ws1,ri,5,row['total'],bg,True,'"S/"#,##0.00')
    for col,w in zip('ABCDE',[24,12,16,10,14]): ws1.column_dimensions[col].width=w

    # Hoja por día
    ws2=wb.create_sheet('Por Día')
    ws2.merge_cells('A1:C1'); t3=ws2['A1']; t3.value='Ventas por día'
    t3.font=Font(bold=True,size=12,color='FFFFFF'); t3.fill=PatternFill('solid',fgColor=verde)
    t3.alignment=Alignment(horizontal='center'); ws2.row_dimensions[1].height=24
    for i,h in enumerate(['Fecha','Pedidos','Total (S/)'],1): hcell(ws2,2,i,h)
    for ri,row in enumerate(por_dia,3):
        bg='E8F5E9' if ri%2==0 else 'FFFFFF'
        dcell(ws2,ri,1,row['fecha'],bg); dcell(ws2,ri,2,row['cantidad'],bg,True)
        dcell(ws2,ri,3,row['total'],bg,True,'"S/"#,##0.00')
    for col,w in zip('ABC',[14,12,14]): ws2.column_dimensions[col].width=w

    # Hoja resumen mesas
    ws3=wb.create_sheet('Por Mesa')
    ws3.merge_cells('A1:C1'); t4=ws3['A1']; t4.value='Resumen por mesa'
    t4.font=Font(bold=True,size=12,color='FFFFFF'); t4.fill=PatternFill('solid',fgColor=naranja)
    t4.alignment=Alignment(horizontal='center'); ws3.row_dimensions[1].height=24
    for i,h in enumerate(['Mesa','Pedidos','Total (S/)'],1): hcell(ws3,2,i,h,bg=naranja)
    for ri,row in enumerate(por_mesa,3):
        bg='FFF3E0' if ri%2==0 else 'FFFFFF'
        dcell(ws3,ri,1,f"Mesa {row['mesa']}",bg); dcell(ws3,ri,2,row['pedidos'],bg,True)
        dcell(ws3,ri,3,row['total'],bg,True,'"S/"#,##0.00')
    for col,w in zip('ABC',[12,12,14]): ws3.column_dimensions[col].width=w

    # Hoja detalle por mesa
    for mesa_row in mesas_unicas:
        mesa_num=mesa_row['mesa']
        ventas_mesa=conn.execute(
            'SELECT * FROM ventas WHERE fecha BETWEEN ? AND ? AND mesa=? ORDER BY fecha,hora',
            (inicio,fin,mesa_num)).fetchall()
        total_mesa=sum(r['precio'] for r in ventas_mesa)
        wsm=wb.create_sheet(f'Mesa {mesa_num}')
        wsm.merge_cells('A1:G1'); tm=wsm['A1']; tm.value=f'Mesa {mesa_num} — S/ {total_mesa:.2f}'
        tm.font=Font(bold=True,size=12,color='FFFFFF'); tm.fill=PatternFill('solid',fgColor=naranja)
        tm.alignment=Alignment(horizontal='center'); wsm.row_dimensions[1].height=24
        for i,h in enumerate(['Fecha','Hora','Categoría','Producto','Tamaño','Personalización','Precio'],1):
            hcell(wsm,2,i,h,bg='424242')
        for ri,row in enumerate(ventas_mesa,3):
            bg='FFF3E0' if ri%2==0 else 'FFFFFF'
            dcell(wsm,ri,1,row['fecha'],bg); dcell(wsm,ri,2,row['hora'],bg)
            dcell(wsm,ri,3,row['categoria'],bg); dcell(wsm,ri,4,row['producto'],bg,align='left')
            dcell(wsm,ri,5,row['tamano'] or '-',bg); dcell(wsm,ri,6,row['personalizado'] or '-',bg,align='left')
            dcell(wsm,ri,7,row['precio'],bg,True,'"S/"#,##0.00')
        for col,w in zip('ABCDEFG',[12,10,14,22,10,26,12]): wsm.column_dimensions[col].width=w

    # Hoja detalle completo
    ws4=wb.create_sheet('Detalle')
    ws4.merge_cells('A1:I1'); t5=ws4['A1']; t5.value='Detalle completo'
    t5.font=Font(bold=True,size=11,color='FFFFFF'); t5.fill=PatternFill('solid',fgColor='424242')
    t5.alignment=Alignment(horizontal='center'); ws4.row_dimensions[1].height=22
    for i,h in enumerate(['Fecha','Hora','Mesa','Categoría','Producto','Tamaño','Personalización','Pago','Precio'],1):
        hcell(ws4,2,i,h,bg='424242')
    for ri,row in enumerate(rows,3):
        bg='F9FBE7' if ri%2==0 else 'FFFFFF'
        vals=[row['fecha'],row['hora'],row['mesa'] or '-',row['categoria'],row['producto'],
              row['tamano'] or '-',row['personalizado'] or '-',row['metodo_pago'],row['precio']]
        for ci,v in enumerate(vals,1):
            dcell(ws4,ri,ci,v,bg,ci==9,'"S/"#,##0.00' if ci==9 else None)
    for col,w in zip('ABCDEFGHI',[12,10,8,14,22,10,26,10,12]): ws4.column_dimensions[col].width=w

    conn.close()
    output=io.BytesIO(); wb.save(output); output.seek(0)
    nombre=f"Reporte_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(output,download_name=nombre,as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
